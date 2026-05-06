#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Отчёт кампании WOW: просмотры по неделям (30 дней), порядок строк — по сценариям API,
а не по дате публикации. HTML + Excel.

Недели считаются от date_from (день 1 = неделя 1): дни 1–7, 8–14, …
skip_weeks — номера недель (1-based), которые исключаются из сумм и помечаются в таблице.
"""

from __future__ import annotations

import argparse
import html
import json
import re
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from src.core.urls_splitter import normalize_url


def platform_label(platform: str) -> str:
    p = (platform or "").strip().lower()
    mapping = {
        "vk": "ВК",
        "tiktok": "ТикТок",
        "youtube": "Ютуб",
        "dzen": "Дзен",
        "pinterest": "Pinterest",
        "ok": "Одноклассники",
    }
    return mapping.get(p, (platform or "").upper() or "—")


def parse_bool_env(val: str | None) -> bool:
    return (val or "").strip().lower() in ("1", "true", "yes", "on")


def parse_date_flexible(s: Any) -> date | None:
    if s is None:
        return None
    if isinstance(s, datetime):
        return s.date()
    if isinstance(s, date):
        return s
    text = str(s).strip()
    if not text:
        return None
    if "T" in text:
        text = text.split("T", 1)[0]
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%Y.%m.%d"):
        try:
            return datetime.strptime(text[:10], fmt).date()
        except ValueError:
            continue
    return None


def load_json(path: Path, default: Any) -> Any:
    if not path.exists():
        return default
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return default


def week_index_for(day: date, start: date, num_weeks: int) -> int | None:
    if day < start:
        return None
    delta = (day - start).days
    idx = delta // 7
    if idx >= num_weeks:
        idx = num_weeks - 1
    return idx


def merge_slot_meta(entries: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    """norm_url -> [meta...] (без схлопывания, чтобы не терять неделю при дублях URL)."""
    out: dict[str, list[dict[str, Any]]] = {}
    for e in entries:
        url = e.get("url") or ""
        if not url:
            continue
        key = normalize_url(url)
        out.setdefault(key, []).append(
            {
                "scenario_id": e.get("scenario_id"),
                "slot_date": e.get("slot_date") or "",
                "blogger_user_id": e.get("blogger_user_id"),
                "platform_type": (e.get("platform_type") or "").strip().lower(),
            }
        )
    return out


def scenario_sort_key(scenarios: list[dict[str, Any]]) -> dict[int, int]:
    order: dict[int, int] = {}
    for i, sc in enumerate(scenarios):
        sid = sc.get("scenario_id")
        if sid is not None:
            try:
                order[int(sid)] = i
            except (TypeError, ValueError):
                continue
    return order


def num_weeks_for_range(start: date, end: date) -> int:
    days = (end - start).days + 1
    return max(1, (days + 6) // 7)


def build_rows(
    unified: list[dict[str, Any]],
    slots_meta: dict[str, list[dict[str, Any]]],
    scenarios: list[dict[str, Any]],
    date_from: date,
    date_to: date,
    skip_weeks_1based: list[int],
    min_weeks: int = 1,
) -> tuple[list[dict[str, Any]], int]:
    """Возвращает строки отчёта и число недель."""
    n_weeks = max(num_weeks_for_range(date_from, date_to), int(min_weeks or 1))
    skip_set = set(skip_weeks_1based)
    scen_order = scenario_sort_key(scenarios)
    scen_title: dict[int, str] = {}
    for sc in scenarios:
        sid = sc.get("scenario_id")
        if sid is None:
            continue
        try:
            scen_title[int(sid)] = str(sc.get("title") or "")
        except (TypeError, ValueError):
            continue

    rows: list[dict[str, Any]] = []

    for item in unified:
        url = item.get("url") or ""
        if not url:
            continue
        key = normalize_url(url)
        pub = parse_date_flexible(item.get("date"))

        views = int(item.get("views", 0) or 0)
        platform = item.get("platform") or ""
        metas = slots_meta.get(key) or [{}]
        # Если по URL несколько слотов, создаём строки по каждому слоту (иначе теряется неделя).
        for meta in metas:
            scenario_id = meta.get("scenario_id")
            slot_d = parse_date_flexible(meta.get("slot_date"))
            # Для календарного отчёта приоритет у slot_date (а не даты публикации из соцсети).
            ref_date = slot_d or pub or date_from

            widx = week_index_for(ref_date, date_from, n_weeks)
            if widx is None:
                widx = 0

            week_views = [0] * n_weeks
            week_human = widx + 1
            if week_human not in skip_set and 0 <= widx < n_weeks:
                week_views[widx] = views

            blogger_name = (
                item.get("owner_name")
                or item.get("author_name")
                or item.get("author")
                or item.get("owner_unique_id")
                or item.get("owner_id")
                or meta.get("blogger_user_id")
                or "—"
            )
            scenario_title = ""
            if scenario_id is not None:
                try:
                    scenario_title = scen_title.get(int(scenario_id), "")
                except (TypeError, ValueError):
                    scenario_title = ""

            rows.append(
                {
                    "scenario_id": scenario_id,
                    "scenario_sort": scen_order.get(int(scenario_id), 10**9) if scenario_id is not None else 10**9 + 1,
                    "scenario_title": scenario_title,
                    "platform": platform,
                    "url": url,
                    "title": (item.get("description") or "")[:120],
                    "ref_date": ref_date.isoformat(),
                    "views_total": views,
                    "week_views": week_views,
                    "week_index_1based": week_human,
                    "blogger_name": str(blogger_name).strip() or "—",
                    "blogger_user_id": meta.get("blogger_user_id"),
                }
            )

    # Требование: сортировка по названиям сценариев по возрастанию.
    rows.sort(
        key=lambda x: (
            (x.get("scenario_title") or "zzzzzzzz").lower(),
            x.get("ref_date") or "",
            (x.get("blogger_name") or "").lower(),
            (x.get("platform") or "").lower(),
            x["url"],
        )
    )

    return rows, n_weeks


def sums_by_week(rows: list[dict[str, Any]], n_weeks: int, skip_weeks_1based: list[int]) -> list[int]:
    skip = set(skip_weeks_1based)
    sums = [0] * n_weeks
    for r in rows:
        for i in range(n_weeks):
            if (i + 1) in skip:
                continue
            sums[i] += r["week_views"][i]
    return sums


def render_html(
    rows: list[dict[str, Any]],
    n_weeks: int,
    skip_weeks: list[int],
    campaign_id: str,
    date_from: str,
    date_to: str,
    sums: list[int],
    validation: dict[str, Any] | None,
    out_path: Path,
) -> None:
    skip_set = set(skip_weeks)
    head_cells = "".join(
        f"<th>Неделя {i + 1}</th>" if (i + 1) not in skip_set else f"<th>Неделя {i + 1} (пропуск)</th>"
        for i in range(n_weeks)
    )
    body = []
    for r in rows:
        wcells = []
        for i in range(n_weeks):
            v = r["week_views"][i]
            if (i + 1) in skip_set:
                wcells.append('<td class="skipped">—</td>')
            else:
                wcells.append(f"<td>{v}</td>")
        sid = r.get("scenario_id")
        sid_s = str(sid) if sid is not None else "—"
        title_s = html.escape(r.get("scenario_title") or "")
        href_esc = html.escape(r["url"], quote=True)
        body.append(
            "<tr>"
            f"<td>{html.escape(sid_s)}</td>"
            f"<td>{title_s}</td>"
            f"<td>{html.escape(r.get('platform') or '')}</td>"
            f"<td class=\"url\"><a href=\"{href_esc}\">{html.escape(r['url'])}</a></td>"
            f"<td>{html.escape(r.get('ref_date') or '')}</td>"
            f"{''.join(wcells)}"
            f"<td>{r['views_total']}</td>"
            "</tr>"
        )

    sum_cells = []
    for i in range(n_weeks):
        if (i + 1) in skip_set:
            sum_cells.append('<td class="skipped">—</td>')
        else:
            sum_cells.append(f"<td>{sums[i]}</td>")

    validation = validation or {}
    duplicates = validation.get("duplicates") if isinstance(validation.get("duplicates"), list) else []
    mismatches = (
        validation.get("platform_mismatches") if isinstance(validation.get("platform_mismatches"), list) else []
    )
    dup_html = "".join(
        f"<tr><td>{int(d.get('count') or 0)}</td><td class='url'>{html.escape(d.get('url_normalized') or '')}</td>"
        f"<td>{html.escape(', '.join(d.get('blogger_ids') or []))}</td></tr>"
        for d in duplicates[:200]
    )
    mismatch_html = "".join(
        "<tr>"
        f"<td>{html.escape(str(m.get('blogger_user_id') or ''))}</td>"
        f"<td>{html.escape(str(m.get('scenario_id') or ''))}</td>"
        f"<td>{html.escape(m.get('expected_platform') or '')}</td>"
        f"<td>{html.escape(m.get('actual_platform') or '')}</td>"
        f"<td class='url'>{html.escape(m.get('url') or '')}</td>"
        "</tr>"
        for m in mismatches[:300]
    )

    html_doc = f"""<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="utf-8"/>
<title>Кампания WOW — недельный отчёт</title>
<style>
body {{ font-family: system-ui, sans-serif; margin: 24px; background: #f8fafc; }}
h1 {{ font-size: 20px; }}
.meta {{ color: #475569; margin-bottom: 16px; }}
table {{ border-collapse: collapse; width: 100%; background: #fff; box-shadow: 0 1px 3px rgba(0,0,0,.08); }}
th, td {{ border: 1px solid #e2e8f0; padding: 8px 10px; font-size: 13px; vertical-align: top; }}
th {{ background: #f1f5f9; text-align: left; }}
tr:nth-child(even) {{ background: #fafafa; }}
td.url {{ max-width: 420px; word-break: break-all; }}
.skipped {{ background: #fff7ed; color: #9a3412; }}
tfoot td {{ font-weight: 600; background: #eef2ff; }}
</style>
</head>
<body>
<h1>Недельный отчёт (порядок по сценариям)</h1>
<div class="meta">Кампания ID: {html.escape(campaign_id)} · период {html.escape(date_from)} — {html.escape(date_to)} · пропуск недель: {html.escape(str(skip_weeks) if skip_weeks else "нет")}</div>
<h2>Проверка ссылок из WOW слотов</h2>
<div class="meta">Дубли: {int(validation.get('duplicates_count') or 0)} · Несоответствие platform/url: {int(validation.get('platform_mismatches_count') or 0)}</div>
<h3>Дубли ссылок</h3>
<table>
<thead><tr><th>Количество</th><th>Нормализованная ссылка</th><th>Blogger IDs</th></tr></thead>
<tbody>{dup_html or '<tr><td colspan="3">Дубли не найдены</td></tr>'}</tbody>
</table>
<h3>Несоответствие соцсети</h3>
<table>
<thead><tr><th>Blogger ID</th><th>Scenario ID</th><th>Ожидалось (platform_type)</th><th>Фактически по URL</th><th>URL</th></tr></thead>
<tbody>{mismatch_html or '<tr><td colspan="5">Несоответствий не найдено</td></tr>'}</tbody>
</table>
<br/>
<table>
<thead>
<tr>
<th>ID сценария</th>
<th>Сценарий</th>
<th>Платформа</th>
<th>URL</th>
<th>Дата (публикация / слот)</th>
{head_cells}
<th>Всего просмотров</th>
</tr>
</thead>
<tbody>
{''.join(body)}
</tbody>
<tfoot>
<tr>
<td colspan="5">Итого по неделям (с учётом пропусков)</td>
{''.join(sum_cells)}
<td>{sum(r['views_total'] for r in rows)}</td>
</tr>
</tfoot>
</table>
</body>
</html>"""
    out_path.write_text(html_doc, encoding="utf-8")


def write_excel(
    rows: list[dict[str, Any]],
    n_weeks: int,
    skip_weeks: list[int],
    campaign_id: str,
    date_from: str,
    date_to: str,
    sums: list[int],
    out_path: Path,
) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
    except ImportError as e:
        raise RuntimeError("Для Excel нужен пакет openpyxl: pip install openpyxl") from e

    skip_set = set(skip_weeks)
    wb = Workbook()
    wb.remove(wb.active)

    date_pairs = 7  # как в референсе: 7 пар "Дата/Ссылка"
    platform_order = ["vk", "tiktok", "youtube", "pinterest"]
    platform_titles = {p: platform_label(p) for p in platform_order}

    for week in range(1, n_weeks + 1):
        ws = wb.create_sheet(f"{week} неделя")

        ws.cell(row=1, column=1, value="Креатор").font = Font(bold=True)
        ws.cell(row=1, column=2, value="Площадка").font = Font(bold=True)
        for i in range(date_pairs):
            c_date = 3 + i * 2
            c_link = c_date + 1
            ws.cell(row=1, column=c_date, value="Дата").font = Font(bold=True)
            ws.cell(row=1, column=c_link, value="Ссылка на публикацию").font = Font(bold=True)

        for col in range(1, 3 + date_pairs * 2):
            ws.cell(row=1, column=col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        week_rows = [r for r in rows if int(r.get("week_index_1based") or 0) == week]
        week_rows.sort(
            key=lambda x: (
                (x.get("blogger_name") or "").lower(),
                (x.get("scenario_title") or "zzzzzzzz").lower(),
                (x.get("platform") or "").lower(),
                x.get("ref_date") or "",
                x.get("url") or "",
            )
        )

        if week in skip_set:
            ws.cell(row=2, column=1, value=f"Неделя {week} помечена как пропуск (skip_weeks)")
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2 + date_pairs * 2)
            ws.cell(row=2, column=1).font = Font(bold=True)
            ws.cell(row=2, column=1).alignment = Alignment(horizontal="left", vertical="center")
            continue

        d0 = parse_date_flexible(date_from)
        week_start = d0 + timedelta(days=(week - 1) * 7) if d0 else None
        day_slots: list[date | None] = []
        for i in range(date_pairs):
            if week_start:
                d = week_start + timedelta(days=i)
                day_slots.append(d if parse_date_flexible(date_to) is None or d <= parse_date_flexible(date_to) else None)
            else:
                day_slots.append(None)

        by_blogger: dict[str, list[dict[str, Any]]] = {}
        for r in week_rows:
            b = r.get("blogger_name") or "—"
            by_blogger.setdefault(str(b), []).append(r)

        row_idx = 2
        for blogger in sorted(by_blogger.keys(), key=lambda x: x.lower()):
            items = by_blogger[blogger]
            by_platform: dict[str, list[dict[str, Any]]] = {p: [] for p in platform_order}
            for it in items:
                p = (it.get("platform") or "").strip().lower()
                if p in by_platform:
                    by_platform[p].append(it)

            blogger_start = row_idx
            # Ровно 4 соцсети * 2 публикации в день = 8 строк на блогера.
            # Слоты публикаций (1/2) определяются по сценариям (по возрастанию названий).
            for platform_key in platform_order:
                platform_items = by_platform.get(platform_key, [])
                scenario_names = sorted(
                    {
                        (x.get("scenario_title") or "").strip() or "—"
                        for x in platform_items
                    },
                    key=lambda s: s.lower(),
                )
                slot_scenarios = (scenario_names + ["—", "—"])[:2]
                for slot_idx in range(2):
                    scenario_name = slot_scenarios[slot_idx]
                    ws.cell(row=row_idx, column=2, value=platform_titles[platform_key])
                    ws.cell(row=row_idx, column=2).alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)

                    for i in range(date_pairs):
                        c_date = 3 + i * 2
                        c_link = c_date + 1
                        ds = day_slots[i]
                        date_txt = ds.strftime("%d.%m") if ds else ""
                        ws.cell(row=row_idx, column=c_date, value=date_txt)
                        ws.cell(row=row_idx, column=c_date).alignment = Alignment(horizontal="center", vertical="top")

                        posts = []
                        for it in platform_items:
                            ref_d = parse_date_flexible(it.get("ref_date"))
                            if not ref_d or not week_start:
                                continue
                            delta = (ref_d - week_start).days
                            if not (0 <= delta < date_pairs) or delta != i:
                                continue
                            title = (it.get("scenario_title") or "").strip() or "—"
                            if title == scenario_name:
                                posts.append(it)
                        if posts:
                            posts.sort(key=lambda x: (x.get("ref_date") or "", x.get("url") or ""))
                            links = [f"{p.get('url') or ''} ({int(p.get('views_total') or 0)})" for p in posts]
                            ws.cell(row=row_idx, column=c_link, value="\n".join(links))
                        else:
                            ws.cell(row=row_idx, column=c_link, value="")
                        ws.cell(row=row_idx, column=c_link).alignment = Alignment(vertical="top", wrap_text=True)
                    row_idx += 1

            blogger_end = row_idx - 1
            ws.cell(row=blogger_start, column=1, value=blogger)
            ws.cell(row=blogger_start, column=1).alignment = Alignment(vertical="top", wrap_text=True)
            if blogger_end > blogger_start:
                ws.merge_cells(start_row=blogger_start, start_column=1, end_row=blogger_end, end_column=1)
                # Визуально как в шаблоне: даты единые для блока блогера.
                for i in range(date_pairs):
                    c_date = 3 + i * 2
                    ws.merge_cells(start_row=blogger_start, start_column=c_date, end_row=blogger_end, end_column=c_date)
                    ds = day_slots[i]
                    ws.cell(row=blogger_start, column=c_date, value=ds.strftime("%d.%m") if ds else "")
                    ws.cell(row=blogger_start, column=c_date).alignment = Alignment(horizontal="center", vertical="top")

        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 14
        for i in range(date_pairs):
            c_date = 3 + i * 2
            c_link = c_date + 1
            letter_date = ws.cell(row=1, column=c_date).column_letter
            letter_link = ws.cell(row=1, column=c_link).column_letter
            ws.column_dimensions[letter_date].width = 10
            ws.column_dimensions[letter_link].width = 34

    meta = wb.create_sheet("Мета", 0)
    meta["A1"] = "Кампания"
    meta["B1"] = campaign_id
    meta["A2"] = "Период"
    meta["B2"] = f"{date_from} — {date_to}"
    meta["A3"] = "Пропуск недель (1-based)"
    meta["B3"] = ", ".join(str(x) for x in sorted(skip_set)) if skip_set else "—"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def generate_weekly_report(work_dir: Path | str) -> tuple[Path, Path] | None:
    work_dir = Path(work_dir)
    ctx_path = work_dir / "wow_campaign_context.json"
    meta_path = work_dir / "wow_slots_meta.json"
    report_path = work_dir / "report.json"

    if not ctx_path.exists() or not meta_path.exists():
        return None

    ctx = load_json(ctx_path, {})
    if not ctx.get("scenarios"):
        pass

    entries = load_json(meta_path, {}).get("entries") or []
    if not isinstance(entries, list):
        entries = []

    unified = load_json(report_path, [])
    if not isinstance(unified, list):
        unified = []

    date_from_s = ctx.get("date_from") or ""
    date_to_s = ctx.get("date_to") or ""
    d0 = parse_date_flexible(date_from_s)
    d1 = parse_date_flexible(date_to_s)
    if not d0 or not d1:
        print("wow_weekly_report: нет корректных date_from/date_to в контексте", flush=True)
        return None

    skip_weeks = ctx.get("skip_weeks") or []
    if not isinstance(skip_weeks, list):
        skip_weeks = []

    campaign_id = str(ctx.get("campaign_id") or "")

    slots_map = merge_slot_meta(entries)
    scenarios = ctx.get("scenarios") or []
    if not isinstance(scenarios, list):
        scenarios = []

    # Для календаря WOW ожидаем недельные вкладки 1..5 (30 дней) даже при кривых датах в контексте.
    # Также гарантируем, что если в skip_weeks указана неделя N, то лист N будет создан.
    required_weeks = max(5, max([int(x) for x in skip_weeks], default=0))
    rows, n_weeks = build_rows(
        unified,
        slots_map,
        scenarios,
        d0,
        d1,
        [int(x) for x in skip_weeks],
        min_weeks=required_weeks,
    )
    sums = sums_by_week(rows, n_weeks, [int(x) for x in skip_weeks])

    html_out = work_dir / "campaign_weekly_report.html"
    xlsx_out = work_dir / "campaign_weekly_report.xlsx"

    validation = ctx.get("slots_validation") if isinstance(ctx.get("slots_validation"), dict) else {}
    render_html(rows, n_weeks, list(skip_weeks), campaign_id, date_from_s, date_to_s, sums, validation, html_out)

    try:
        write_excel(rows, n_weeks, list(skip_weeks), campaign_id, date_from_s, date_to_s, sums, xlsx_out)
    except RuntimeError as exc:
        print(str(exc), flush=True)
        xlsx_out = None  # type: ignore

    print(f"Недельный отчёт HTML: {html_out}", flush=True)
    if xlsx_out:
        print(f"Недельный отчёт Excel: {xlsx_out}", flush=True)

    return html_out, xlsx_out or html_out


def main():
    parser = argparse.ArgumentParser(description="Генерация недельного отчёта WOW (HTML + Excel)")
    parser.add_argument("--work-dir", default=".", help="Каталог с wow_campaign_context.json, wow_slots_meta.json, report.json")
    args = parser.parse_args()
    generate_weekly_report(args.work_dir)


if __name__ == "__main__":
    main()
